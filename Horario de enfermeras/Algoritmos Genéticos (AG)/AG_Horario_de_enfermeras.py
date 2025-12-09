import numpy as np
import pandas as pd
import random
import matplotlib.pyplot as plt
import seaborn as sns
from deap import base, creator, tools, algorithms
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# ==============================================================================
# CLASE: PROBLEMA DE PROGRAMACIÓN DE ENFERMERAS (Antiguo nurses.py)
# ==============================================================================

class ProblemaProgramacionEnfermeras:
    """Esta clase encapsula el problema de Programación de Horarios de Enfermeras"""

    def __init__(self, penalizacion_restriccion_dura):
        """
        :param penalizacion_restriccion_dura: factor de penalización por violar una restricción dura
        """
        self.penalizacion_restriccion_dura = penalizacion_restriccion_dura

        # Lista de enfermeras:
        self.enfermeras = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

        # Preferencias de turno de cada enfermera - [Mañana, Tarde, Noche]:
        # 1 = Prefiere ese turno, 0 = No lo prefiere
        self.preferencia_turnos = [
            [1, 0, 0], [1, 1, 0], [0, 0, 1], [0, 1, 0], 
            [0, 0, 1], [1, 1, 1], [0, 1, 1], [1, 1, 1]
        ]

        # Mínimo y máximo de enfermeras permitidas por turno - [Mañana, Tarde, Noche]:
        self.min_turno = [2, 2, 1]
        self.max_turno = [3, 4, 2]

        # Máximo de turnos por semana permitidos para cada enfermera:
        self.max_turnos_por_semana = 5

        # Número de semanas para las que creamos el horario:
        self.semanas = 1

        # Valores útiles derivados:
        self.turnos_por_dia = len(self.min_turno)
        self.turnos_por_semana = 7 * self.turnos_por_dia

    def __len__(self):
        """
        :return: el número total de turnos en el horario (cromosoma)
        """
        return len(self.enfermeras) * self.turnos_por_semana * self.semanas

    def obtener_costo(self, horario):
        """
        Calcula el costo total (fitness) basado en las violaciones del horario dado.
        :param horario: lista de valores binarios que describe el horario.
        :return: el costo calculado.
        """
        if len(horario) != self.__len__():
            raise ValueError("El tamaño de la lista de horario debe ser igual a ", self.__len__())

        # Convertir el horario lineal en un diccionario por enfermera:
        dict_turnos_enfermera = self.obtener_turnos_por_enfermera(horario)

        # Contar las diversas violaciones:
        violaciones_consecutivas = self.contar_violaciones_turnos_consecutivos(dict_turnos_enfermera)
        violaciones_turnos_semana = self.contar_violaciones_turnos_por_semana(dict_turnos_enfermera)[1]
        violaciones_enfermeras_turno = self.contar_violaciones_enfermeras_por_turno(dict_turnos_enfermera)[1]
        violaciones_preferencia = self.contar_violaciones_preferencia(dict_turnos_enfermera)

        # Calcular el costo de las violaciones:
        violaciones_duras = violaciones_consecutivas + violaciones_enfermeras_turno + violaciones_turnos_semana
        violaciones_suaves = violaciones_preferencia

        return self.penalizacion_restriccion_dura * violaciones_duras + violaciones_suaves

    def obtener_turnos_por_enfermera(self, horario):
        """
        Convierte el horario completo en un diccionario separado por enfermera.
        :return: diccionario con la enfermera como clave y sus turnos como valor.
        """
        turnos_por_enfermera = self.__len__() // len(self.enfermeras)
        dict_turnos = {}
        indice_turno = 0

        for enfermera in self.enfermeras:
            dict_turnos[enfermera] = horario[indice_turno:indice_turno + turnos_por_enfermera]
            indice_turno += turnos_por_enfermera

        return dict_turnos

    def contar_violaciones_turnos_consecutivos(self, dict_turnos_enfermera):
        """
        Cuenta violaciones si una enfermera trabaja turnos consecutivos (ej. mañana y tarde el mismo día).
        """
        violaciones = 0
        for turnos in dict_turnos_enfermera.values():
            # Busca dos '1' consecutivos:
            for turno1, turno2 in zip(turnos, turnos[1:]):
                if turno1 == 1 and turno2 == 1:
                    violaciones += 1
        return violaciones

    def contar_violaciones_turnos_por_semana(self, dict_turnos_enfermera):
        """
        Cuenta violaciones si se excede el máximo de turnos por semana.
        """
        violaciones = 0
        lista_turnos_semanales = []
        
        for turnos in dict_turnos_enfermera.values():
            for i in range(0, self.semanas * self.turnos_por_semana, self.turnos_por_semana):
                turnos_semanales = sum(turnos[i:i + self.turnos_por_semana])
                lista_turnos_semanales.append(turnos_semanales)
                if turnos_semanales > self.max_turnos_por_semana:
                    violaciones += turnos_semanales - self.max_turnos_por_semana

        return lista_turnos_semanales, violaciones

    def contar_violaciones_enfermeras_por_turno(self, dict_turnos_enfermera):
        """
        Cuenta violaciones si hay muchas o pocas enfermeras en un turno específico.
        """
        total_por_turno = [sum(turno) for turno in zip(*dict_turnos_enfermera.values())]
        violaciones = 0
        
        for indice, num_enfermeras in enumerate(total_por_turno):
            indice_diario = indice % self.turnos_por_dia  # 0, 1 o 2
            if num_enfermeras > self.max_turno[indice_diario]:
                violaciones += num_enfermeras - self.max_turno[indice_diario]
            elif num_enfermeras < self.min_turno[indice_diario]:
                violaciones += self.min_turno[indice_diario] - num_enfermeras

        return total_por_turno, violaciones

    def contar_violaciones_preferencia(self, dict_turnos_enfermera):
        """
        Cuenta violaciones a las preferencias de turno de las enfermeras.
        """
        violaciones = 0
        for indice, preferencia in enumerate(self.preferencia_turnos):
            # Duplicar la preferencia para cubrir toda la semana
            pref_completa = preferencia * (self.turnos_por_semana // self.turnos_por_dia)
            turnos = dict_turnos_enfermera[self.enfermeras[indice]]
            
            for pref, turno_real in zip(pref_completa, turnos):
                if pref == 0 and turno_real == 1:
                    violaciones += 1
        return violaciones

    def imprimir_info_horario(self, horario):
        """
        Muestra la información del horario en consola y exporta a Excel.
        """
        dict_turnos_enfermera = self.obtener_turnos_por_enfermera(horario)

        # --- Días y turnos ---
        dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        turnos_dict = {
            (1, 0, 0): "Mañana",
            (0, 1, 0): "Tarde",
            (0, 0, 1): "Noche",
            (0, 0, 0): "Descanso",
            (1, 1, 0): "Mañana-Tarde",
            (1, 0, 1): "Mañana-Noche",
            (0, 1, 1): "Tarde-Noche",
            (1, 1, 1): "Triple turno"
        }

        # --- Construir tabla de turnos ---
        tabla = {}
        for enfermera, turnos in dict_turnos_enfermera.items():
            dias_turnos = []
            for i in range(0, len(turnos), self.turnos_por_dia):
                bits = tuple(turnos[i:i + self.turnos_por_dia])
                dias_turnos.append(turnos_dict.get(bits, "Error"))
            tabla[enfermera] = dias_turnos

        df = pd.DataFrame(tabla, index=dias_semana)
        print("\n📅 Horario semanal ideal (turnos por día y enfermera):\n")
        print(df.T)
        print()

        # --- Exportar a Excel ---
        self.exportar_horario_excel(df.T)

        # --- Mostrar información de violaciones ---
        print("➡ Violaciones de turnos consecutivos =", self.contar_violaciones_turnos_consecutivos(dict_turnos_enfermera))
        print()

        lista_semanal, violaciones = self.contar_violaciones_turnos_por_semana(dict_turnos_enfermera)
        print("➡ Turnos por semana (por enfermera) =", lista_semanal)
        print("➡ Violaciones de máx. turnos por semana =", violaciones)
        print()

        total_por_turno, violaciones = self.contar_violaciones_enfermeras_por_turno(dict_turnos_enfermera)
        print("➡ Enfermeras por turno (total) =", total_por_turno)
        print("➡ Violaciones de cantidad de enfermeras =", violaciones)
        print()

        violaciones_pref = self.contar_violaciones_preferencia(dict_turnos_enfermera)
        print("➡ Violaciones de preferencia de turno =", violaciones_pref)
        print()

    def exportar_horario_excel(self, df, nombre_archivo="horario_enfermeras.xlsx"):
        """
        Exporta el horario a un archivo Excel con formato visual.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Horario"

        # --- Colores vivos ---
        colores = {
            "Mañana": "FFF176",
            "Tarde": "4FC3F7",
            "Noche": "BA68C8",
            "Descanso": "BDBDBD",
            "Mañana-Tarde": "81C784",
            "Tarde-Noche": "FF8A65",
            "Mañana-Noche": "F06292",
            "Triple turno": "FFD54F",
        }

        borde = Border(left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium"))
        alineado = Alignment(horizontal="center", vertical="center")

        # --- Encabezado ---
        ws.append(["Enfermera"] + list(df.columns))
        for col in range(1, ws.max_column + 1):
            celda = ws.cell(row=1, column=col)
            celda.font = Font(bold=True)
            celda.fill = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
            celda.border = borde
            celda.alignment = alineado

        # --- Insertar datos ---
        for enfermera, row in df.iterrows():
            ws.append([enfermera] + list(row))

        # --- Aplicar formato ---
        for fila in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for celda in fila:
                turno = celda.value
                if turno in colores:
                    celda.fill = PatternFill(start_color=colores[turno], end_color=colores[turno], fill_type="solid")
                celda.border = borde
                celda.alignment = alineado

        # Ajuste de dimensiones
        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = 30
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        wb.save(nombre_archivo)
        print(f"\n📁 Archivo Excel generado correctamente: {nombre_archivo}\n")


# ==============================================================================
# ALGORITMO: ELITISMO (Antiguo elitism.py)
# ==============================================================================

def eaSimpleConElitismo(poblacion, toolbox, cxpb, mutpb, ngen, stats=None,
                        halloffame=None, verbose=__debug__):
    """
    Este algoritmo es similar al eaSimple() de DEAP, con la modificación de que
    el 'halloffame' (Salón de la Fama) se usa para implementar elitismo.
    Los mejores individuos se inyectan directamente en la siguiente generación.
    """
    logbook = tools.Logbook()
    logbook.header = ['gen', 'nevals'] + (stats.fields if stats else [])

    # Evaluar individuos con fitness inválido
    invalid_ind = [ind for ind in poblacion if not ind.fitness.valid]
    fitnesses = toolbox.map(toolbox.evaluate, invalid_ind)
    for ind, fit in zip(invalid_ind, fitnesses):
        ind.fitness.values = fit

    if halloffame is None:
        raise ValueError("¡El parámetro halloffame no debe estar vacío para usar elitismo!")

    halloffame.update(poblacion)
    hof_size = len(halloffame.items) if halloffame.items else 0

    record = stats.compile(poblacion) if stats else {}
    logbook.record(gen=0, nevals=len(invalid_ind), **record)
    if verbose:
        print(logbook.stream)

    # Inicio del proceso generacional
    for gen in range(1, ngen + 1):
        # Seleccionar la siguiente generación (menos los espacios reservados para elitismo)
        offspring = toolbox.select(poblacion, len(poblacion) - hof_size)

        # Variar el pool de individuos (Cruce y Mutación)
        offspring = algorithms.varAnd(offspring, toolbox, cxpb, mutpb)

        # Evaluar individuos con fitness inválido
        invalid_ind = [ind for ind in offspring if not ind.fitness.valid]
        fitnesses = toolbox.map(toolbox.evaluate, invalid_ind)
        for ind, fit in zip(invalid_ind, fitnesses):
            ind.fitness.values = fit

        # --- ELITISMO: Añadir los mejores de vuelta a la población ---
        offspring.extend(halloffame.items)

        # Actualizar el Hall of Fame con los nuevos hijos
        halloffame.update(offspring)

        # Reemplazar la población actual con la descendencia
        poblacion[:] = offspring

        # Registrar estadísticas
        record = stats.compile(poblacion) if stats else {}
        logbook.record(gen=gen, nevals=len(invalid_ind), **record)
        if verbose:
            print(logbook.stream)

    return poblacion, logbook


# ==============================================================================
# MAIN: EJECUCIÓN (Antiguo nurse_scheduling.py)
# ==============================================================================

# Constantes del problema:
PENALIZACION_DURA = 10  # Penalización por violar restricciones duras

# Constantes del Algoritmo Genético:
TAMANO_POBLACION = 300
PROB_CRUCE = 0.9       # Probabilidad de cruce (Crossover)
PROB_MUTACION = 0.1    # Probabilidad de mutación
MAX_GENERACIONES = 200
TAMANO_SALON_FAMA = 30 # Cuántos mejores individuos guardamos (Elitismo)

# Semilla aleatoria para reproducibilidad
SEMILLA_RANDOM = 42
random.seed(SEMILLA_RANDOM)

toolbox = base.Toolbox()

# Crear instancia del problema
problema = ProblemaProgramacionEnfermeras(PENALIZACION_DURA)

# Definir estrategia de minimización de fitness (queremos minimizar costo/violaciones)
creator.create("FitnessMin", base.Fitness, weights=(-1.0,))

# Crear clase Individuo
creator.create("Individual", list, fitness=creator.FitnessMin)

# Operador para generar 0 o 1 aleatoriamente
toolbox.register("zeroOrOne", random.randint, 0, 1)

# Operador para crear un individuo (cromosoma)
toolbox.register("individualCreator", tools.initRepeat, creator.Individual, toolbox.zeroOrOne, len(problema))

# Operador para crear la población
toolbox.register("populationCreator", tools.initRepeat, list, toolbox.individualCreator)

# Función de evaluación de fitness
def obtener_costo(individual):
    return problema.obtener_costo(individual),  # Retornar como tupla

toolbox.register("evaluate", obtener_costo)

# Operadores genéticos:
toolbox.register("select", tools.selTournament, tournsize=2)
toolbox.register("mate", tools.cxTwoPoint)
toolbox.register("mutate", tools.mutFlipBit, indpb=1.0/len(problema))

def main():
    print("--- Iniciando Optimización de Horarios de Enfermeras ---")
    
    # Crear población inicial
    poblacion = toolbox.populationCreator(n=TAMANO_POBLACION)

    # Preparar estadísticas
    stats = tools.Statistics(lambda ind: ind.fitness.values)
    stats.register("min", np.min)
    stats.register("promedio", np.mean)

    # Definir el Hall of Fame (Mejores individuos)
    hof = tools.HallOfFame(TAMANO_SALON_FAMA)

    # Ejecutar Algoritmo Genético con Elitismo
    poblacion, logbook = eaSimpleConElitismo(
        poblacion, toolbox,
        cxpb=PROB_CRUCE,
        mutpb=PROB_MUTACION,
        ngen=MAX_GENERACIONES,
        stats=stats,
        halloffame=hof,
        verbose=True
    )

    # --- Resultados ---
    mejor_individuo = hof.items[0]
    print("\n-- Mejor individuo encontrado (Genes) --")
    # print(mejor_individuo) # Descomentar si quieres ver la lista de 0s y 1s
    print("-- Fitness (Costo Total) = ", mejor_individuo.fitness.values[0])

    print("\n-- Horario Óptimo Generado --")
    problema.imprimir_info_horario(mejor_individuo)

    # --- Extraer datos para graficar ---
    valores_min = logbook.select("min")
    valores_promedio = logbook.select("promedio")

    # --- Graficar ---
    sns.set_style("whitegrid")
    plt.figure(figsize=(10, 6))

    generaciones = range(len(valores_min))

    plt.plot(generaciones, valores_min, color='red', linestyle='-', linewidth=2, label='Mejor Fitness (Mínimo Costo)')
    plt.plot(generaciones, valores_promedio, color='green', linestyle='--', linewidth=2, label='Fitness Promedio')

    # Etiquetas del valor final
    plt.text(generaciones[-1], valores_min[-1],
             f"Min: {valores_min[-1]:.2f}",
             color='red', fontsize=10, ha='left', va='bottom')
    
    plt.title('📉 Evolución de la Optimización (Fitness por Generación)', fontsize=13, fontweight='bold')
    plt.xlabel('Generación', fontsize=11)
    plt.ylabel('Costo (Fitness)', fontsize=11)
    plt.legend(loc='upper right', fontsize=10)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.tight_layout()
    
    print("Mostrando gráfica de evolución...")
    plt.show()

if __name__ == "__main__":
    main()